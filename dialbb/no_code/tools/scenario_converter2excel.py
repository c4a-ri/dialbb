# scenario_converter2excel.py
# Export JSON (state_graph.json) -> scenario.xlsx converter (round-trip friendly)

from __future__ import annotations

import argparse
import json
import sys
from collections import deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


NODE_TYPE_SYSTEM = "system"
NODE_TYPE_USER = "user"

HEADERS = [
    "flag",
    "state",
    "system utterance",
    "user utterance example",
    "user utterance type",
    "conditions",
    "actions",
    "next state",
]


@dataclass(frozen=True)
class Node:
    node_id: str
    node_type: str
    text: str
    fields: Dict[str, Any]

    def get(self, key: str, default: str = "") -> str:
        v = self.fields.get(key, default)
        if v is None:
            return default
        return str(v)


@dataclass(frozen=True)
class Edge:
    from_id: str
    to_id: str


def die(msg: str) -> None:
    print(f"ERROR: {msg}")
    sys.exit(1)


def load_graph(json_path: Path) -> Tuple[Dict[str, Node], List[Edge]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))

    nodes: Dict[str, Node] = {}
    for n in data.get("nodes", []):
        node_id = n.get("id", "")
        node_type = n.get("type", "")
        text = n.get("text", "")
        fields = dict(n)
        for k in ["id", "x", "y", "text", "type", "short_id"]:
            fields.pop(k, None)

        if node_id:
            nodes[node_id] = Node(node_id=node_id, node_type=node_type, text=text, fields=fields)

    edges: List[Edge] = []
    for e in data.get("edges", []):
        a = e.get("from")
        b = e.get("to")
        if a and b:
            edges.append(Edge(from_id=a, to_id=b))

    return nodes, edges


def build_outgoing(nodes: Dict[str, Node], edges: List[Edge]) -> Dict[str, List[str]]:
    outgoing: Dict[str, List[str]] = {nid: [] for nid in nodes.keys()}
    for e in edges:
        if e.from_id in outgoing:
            outgoing[e.from_id].append(e.to_id)
    return outgoing


def system_utterance_value(sys_node: Node) -> str:
    return sys_node.get("utterance", "")


def user_priority_value(user_node: Node) -> int:
    try:
        return int(user_node.get("priority", "0") or 0)
    except ValueError:
        return 0


def user_fields(user_node: Node) -> Tuple[str, str, str, str]:
    # (example, type, conditions, actions)  actions列は現状空だが、将来のために読む
    ex = user_node.get("utterance_example", "")
    ut = user_node.get("utterance_type", "")
    cond = user_node.get("condition", "")
    act = user_node.get("action", "")
    return ex, ut, cond, act


def build_system_state_mapper(nodes: Dict[str, Node], start_system_id: str, outgoing: Dict[str, List[str]]) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    要件の「元CSVに復元できる」ための state 命名。
    - initial/prep/error は固定 (#initial等)
    - final は到達順に #final_stateN
    - other/skip は到達順に stateN
    - skip の utterance は $skip 強制
    """
    system_ids = [nid for nid, n in nodes.items() if n.node_type == NODE_TYPE_SYSTEM]

    # reachability traversal from initial (system -> users -> system)
    q = deque([start_system_id])
    visited_system: set[str] = set()

    # assigners
    state_counter = 0
    final_counter = 0
    system_state: Dict[str, str] = {}
    system_utt: Dict[str, str] = {}

    def assign_system(sys_id: str) -> None:
        nonlocal state_counter, final_counter
        if sys_id in system_state:
            return

        sys_node = nodes[sys_id]
        nk = (sys_node.get("node_kind", "") or "").strip().lower()

        if nk in ("initial", "prep", "error"):
            system_state[sys_id] = f"#{nk}"
            system_utt[sys_id] = system_utterance_value(sys_node)

        elif nk == "final":
            final_counter += 1
            system_state[sys_id] = f"#final_state{final_counter}"
            system_utt[sys_id] = system_utterance_value(sys_node)

        else:  # other
            state_counter += 1
            system_state[sys_id] = f"state{state_counter}"
            system_utt[sys_id] = system_utterance_value(sys_node)

    # seed assign
    assign_system(start_system_id)

    while q:
        cur = q.popleft()
        if cur in visited_system:
            continue
        visited_system.add(cur)

        # system -> users
        user_ids = [tid for tid in outgoing.get(cur, []) if tid in nodes and nodes[tid].node_type == NODE_TYPE_USER]
        # sort users by priority desc (same system block)
        user_ids.sort(key=lambda uid: user_priority_value(nodes[uid]), reverse=True)

        for uid in user_ids:
            # user -> next system (A型チェック)
            next_sys_ids = [tid for tid in outgoing.get(uid, []) if tid in nodes and nodes[tid].node_type == NODE_TYPE_SYSTEM]
            if len(next_sys_ids) > 1:
                die(f"A型のみ許容ですが、User({uid}) が複数Systemへ遷移しています: {next_sys_ids}")
            if len(next_sys_ids) == 1:
                ns = next_sys_ids[0]
                assign_system(ns)
                if ns not in visited_system:
                    q.append(ns)

    # also assign any remaining systems not reached (ex: isolated #error)
    # but keep numbering stable: assign final/other/skip for these *after* traversal
    # (so reached ones keep expected numbers)
    for sid in system_ids:
        if sid not in system_state:
            assign_system(sid)

    return system_state, system_utt


def generate_rows(nodes: Dict[str, Node], edges: List[Edge]) -> List[List[str]]:
    outgoing = build_outgoing(nodes, edges)

    # find unique initial system
    initial_systems = [
        n.node_id for n in nodes.values()
        if n.node_type == NODE_TYPE_SYSTEM and (n.get("node_kind", "") or "").strip().lower() == "initial"
    ]
    if not initial_systems:
        die("node_kind='initial' のSystemノードが見つかりません。")
    if len(initial_systems) > 1:
        die("node_kind='initial' のSystemノードが複数あります（禁止）。")

    start_system_id = initial_systems[0]

    # build mapping
    system_state, system_utt = build_system_state_mapper(nodes, start_system_id, outgoing)

    # traverse again to emit rows in the same “scenario order”
    rows: List[List[str]] = []
    q = deque([start_system_id])
    emitted_system: set[str] = set()

    def emit_system_block(sys_id: str) -> None:
        sys_node = nodes[sys_id]
        state_label = system_state[sys_id]

        # system -> users
        user_ids = [tid for tid in outgoing.get(sys_id, []) if tid in nodes and nodes[tid].node_type == NODE_TYPE_USER]
        user_ids.sort(key=lambda uid: user_priority_value(nodes[uid]), reverse=True)

        if not user_ids:
            # standalone system row
            rows.append(["", state_label, system_utt.get(sys_id, ""), "", "", "", "", ""])
            return

        first = True
        for uid in user_ids:
            u = nodes[uid]
            ex, ut, cond, act = user_fields(u)

            # user -> next system (A型)
            next_sys_ids = [tid for tid in outgoing.get(uid, []) if tid in nodes and nodes[tid].node_type == NODE_TYPE_SYSTEM]
            if len(next_sys_ids) > 1:
                die(f"User({uid}) が複数Systemへ遷移しています: {next_sys_ids}")

            next_state = system_state[next_sys_ids[0]] if len(next_sys_ids) == 1 else ""

            # system utterance is shown only for the first row of this system block
            sys_utt_cell = system_utt.get(sys_id, "") if first else ""
            first = False

            rows.append(["", state_label, sys_utt_cell, ex, ut, cond, act, next_state])

            # enqueue next system (encounter order)
            if len(next_sys_ids) == 1:
                ns = next_sys_ids[0]
                if ns not in emitted_system:
                    q.append(ns)

    while q:
        cur = q.popleft()
        if cur in emitted_system:
            continue
        emitted_system.add(cur)
        emit_system_block(cur)

    # emit remaining systems not reached (e.g. isolated #error)
    remaining = [n.node_id for n in nodes.values() if n.node_type == NODE_TYPE_SYSTEM and n.node_id not in emitted_system]

    # put error at the end to match your sample CSV style
    def rem_key(sid: str) -> Tuple[int, str]:
        nk = (nodes[sid].get("node_kind", "") or "").strip().lower()
        if nk == "error":
            return (99, sid)
        return (0, sid)

    for sid in sorted(remaining, key=rem_key):
        emitted_system.add(sid)
        # standalone row
        rows.append(["", system_state[sid], system_utt.get(sid, ""), "", "", "", "", ""])

    return rows


def write_xlsx(rows: List[List[str]], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "scenario"

    ws.append(HEADERS)
    for r in rows:
        ws.append(r)

    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ------------------------------------------------------------
# GUI用エントリポイント
# ------------------------------------------------------------
def convert_json_to_excel(json_path: str, excel_path: str) -> None:
    """
    GUIから直接呼び出すためのAPI。
    JSON→Excel変換を実行する。
    """
    nodes, edges = load_graph(Path(json_path))
    rows = generate_rows(nodes, edges)
    write_xlsx(rows, Path(excel_path))
    # print(f"OK: {Path(excel_path).resolve()} (rows={len(rows)})")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Convert state_graph.json to scenario.xlsx (round-trip)")
    ap.add_argument("json_path", nargs="?", default="state_graph.json")
    ap.add_argument("-o", "--out", default="scenario.xlsx")
    args = ap.parse_args()

    convert_json_to_excel(args.json_path, args.out)
